"""GEOS — Report Export (PDF + Excel).

Generates official reports with SNN data, projections, and entity details.
Uses fpdf2 for PDF (no PIL dependency) and openpyxl for Excel.
"""

from __future__ import annotations

import io
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── PDF ────────────────────────────────────────────────────────

class GEOSReport(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(26, 82, 118)
        self.cell(0, 8, "GEOS - Government Economic Optimization System", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(26, 82, 118)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128)
        self.cell(0, 10, "Page %d/{nb} - %s" % (self.page_no(), datetime.now().strftime('%d/%m/%Y')), align="C")

    def section_title(self, title: str) -> None:
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(26, 82, 118)
        self.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
        self.ln(2)

    def table(self, headers: list[str], rows: list[list[str]], col_widths: list[float] | None = None) -> None:
        if col_widths is None:
            col_widths = [190 / len(headers)] * len(headers)
        # Header
        self.set_font("Helvetica", "B", 9)
        self.set_fill_color(26, 82, 118)
        self.set_text_color(255)
        for i, h in enumerate(headers):
            self.cell(col_widths[i], 7, h, border=1, fill=True, align="C")
        self.ln()
        # Rows
        self.set_font("Helvetica", "", 8)
        self.set_text_color(0)
        for ri, row in enumerate(rows):
            fill = ri % 2 == 1
            if fill:
                self.set_fill_color(242, 243, 244)
            for i, val in enumerate(row):
                align = "R" if i > 0 and any(c.isdigit() for c in val) else "L"
                self.cell(col_widths[i], 6, str(val), border=1, fill=fill, align=align)
            self.ln()
        self.ln(4)


def generate_snn_pdf(engine, predictions: dict | None = None) -> bytes:
    """Generate a PDF report of the GEOS SNN analysis."""
    pdf = GEOSReport()
    pdf.alias_nb_pages()
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(26, 82, 118)
    pdf.cell(0, 12, "Rapport SNN - GEOS", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(80)
    pdf.cell(0, 8, f"Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, "Formule: max SNN = CS + PS + GR + NRV - DWL - EC", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    # SNN Summary
    agg = engine.compute_snn()
    pdf.section_title("1. Résumé SNN")

    def _pct(v):
        return f"{v / agg.snn * 100:.1f}%" if agg.snn else "0%"

    data = [
        ["CS (Consumer Surplus)", f"{agg.total_cs:,.1f}", _pct(agg.total_cs)],
        ["PS (Producer Surplus)", f"{agg.total_ps:,.1f}", _pct(agg.total_ps)],
        ["GR (Government Revenue)", f"{agg.total_gr:,.1f}", _pct(agg.total_gr)],
        ["NRV (Net Resource Value)", f"{agg.total_nrv:,.1f}", _pct(agg.total_nrv)],
        ["DWL (Deadweight Loss)", f"({agg.total_dwl:,.1f})", f"-{_pct(agg.total_dwl)}"],
        ["EC (Environmental Cost)", f"({agg.total_ec:,.1f})", f"-{_pct(agg.total_ec)}"],
        ["SNN TOTAL", f"{agg.snn:,.1f}", "100%"],
    ]
    pdf.table(["Composant", "Valeur (M USD)", "Contribution"], data, [70, 50, 50])

    # Entities
    pdf.section_title("2. Entités GEOS")
    entity_data = [
        ["Provinces", str(len(engine.provinces)),
         ", ".join(p["name"] for p in engine.provinces[:3])],
        ["Entreprises", str(len(engine.companies)),
         ", ".join(c["name"] for c in engine.companies[:3])],
        ["Ressources", str(len(engine.resources)),
         ", ".join(r["name"] for r in engine.resources[:3])],
        ["Ministères", str(len(engine.ministries)),
         ", ".join(m["name"] for m in engine.ministries[:3])],
        ["Impôts", str(len(engine.taxes)),
         ", ".join(t["name"] for t in engine.taxes[:3])],
        ["Projets", str(len(engine.projects)),
         ", ".join(p["name"] for p in engine.projects[:3])],
        ["Indicateurs", str(len(engine.indicators)), "Macro, Budget, Fiscal"],
    ]
    pdf.table(["Entité", "Nombre", "Exemples"], entity_data, [40, 20, 110])

    # Predictions
    if predictions and "projections" in predictions:
        pdf.add_page()
        pdf.section_title("3. Predictions - Scenario")
        proj_data = []
        for p in predictions["projections"][:10]:
            proj_data.append([
                str(p["year"]),
                f"{p['snn']:,.0f}",
                f"{p['cs']:,.0f}",
                f"{p['ps']:,.0f}",
                f"{p['nrv']:,.0f}",
            ])
        pdf.table(["Année", "SNN", "CS", "PS", "NRV"], proj_data, [25, 40, 40, 40, 40])

        if "mean_final_snn" in predictions:
            pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, f"SNN final moyen (Monte Carlo): {predictions['mean_final_snn']:,.0f} M USD",
                 new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Intervalle 95%: [{predictions['ci_5']:,.0f} - {predictions['ci_95']:,.0f}] M USD",
                 new_x="LMARGIN", new_y="NEXT")

    buf = io.BytesIO()
    pdf_buf = pdf.output()
    if isinstance(pdf_buf, str):
        pdf_buf = pdf_buf.encode("latin-1")
    return bytes(pdf_buf)


# ── Excel ──────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def generate_snn_excel(engine, predictions: dict | None = None) -> bytes:
    """Generate an Excel workbook with all GEOS data."""
    wb = Workbook()

    # Sheet 1: SNN Summary
    ws = wb.active
    ws.title = "SNN Summary"
    agg = engine.compute_snn()

    ws.append(["GEOS — Surplus National Net Analysis"])
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=14, color="1A5276")
    ws.append([])
    ws.append(["Date", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws.append(["Formule", "SNN = CS + PS + GR + NRV - DWL - EC"])
    ws.append([])

    ws.append(["Composant", "Valeur (M USD)", "Contribution %"])
    _style_header(ws, 6, 3)
    components = [
        ("CS", agg.total_cs), ("PS", agg.total_ps),
        ("GR", agg.total_gr), ("NRV", agg.total_nrv),
        ("DWL", -agg.total_dwl), ("EC", -agg.total_ec),
    ]
    for name, val in components:
        pct = round(val / agg.snn * 100, 1) if agg.snn else 0
        ws.append([name, round(val, 2), pct])
    ws.append(["SNN TOTAL", round(agg.snn, 2), 100])

    for col in ["A", "B", "C"]:
        ws.column_dimensions[col].width = 22

    # Sheet 2: Provinces
    ws2 = wb.create_sheet("Provinces")
    ws2.append(["Province", "Population", "Area km²", "Literacy %", "Internet %", "Security"])
    _style_header(ws2, 1, 6)
    for p in engine.provinces:
        ws2.append([p["name"], p.get("population", 0), p.get("area_km2", 0),
                     p.get("literacy_rate", 0), p.get("internet_pct", 0),
                     p.get("security_index", 0)])
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 18

    # Sheet 3: Companies
    ws3 = wb.create_sheet("Entreprises")
    ws3.append(["Entreprise", "Secteur", "Revenue M$", "Cost Total", "PS"])
    _style_header(ws3, 1, 5)
    for c in engine.companies:
        tc = (c.get("production_cost", 0) + c.get("tax_burden", 0)
              + c.get("admin_cost", 0) + c.get("corruption_cost", 0)
              + c.get("logistics_cost", 0) + c.get("energy_cost", 0))
        ps = max(0, c.get("revenue", 0) - tc)
        ws3.append([c["name"], c.get("sector", ""), c.get("revenue", 0),
                     tc, round(ps, 2)])
    for col in ["A", "B", "C", "D", "E"]:
        ws3.column_dimensions[col].width = 22

    # Sheet 4: Resources
    ws4 = wb.create_sheet("Ressources")
    ws4.append(["Ressource", "Type", "Production t", "Value $/t", "Local %", "NRV M$", "EC"])
    _style_header(ws4, 1, 7)
    for r in engine.resources:
        gv = r.get("annual_production_tons", 0) * r.get("market_value_per_ton", 0) / 1_000_000
        nrv = gv * (1 + r.get("local_processing_pct", 0) / 100)
        ws4.append([r["name"], r.get("type", ""), r.get("annual_production_tons", 0),
                     r.get("market_value_per_ton", 0), r.get("local_processing_pct", 0),
                     round(nrv, 2), r.get("environmental_cost", 0)])
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws4.column_dimensions[col].width = 18

    # Sheet 5: Indicators
    ws5 = wb.create_sheet("Indicateurs")
    ws5.append(["Indicateur", "Catégorie", "Valeur", "Unité", "Année", "Source", "Cible"])
    _style_header(ws5, 1, 7)
    for ind in engine.indicators:
        ws5.append([ind["name"], ind.get("category", ""), ind.get("value", 0),
                     ind.get("unit", ""), ind.get("year", ""), ind.get("source", ""),
                     ind.get("target", "")])
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws5.column_dimensions[col].width = 20

    # Sheet 6: Predictions
    if predictions and "projections" in predictions:
        ws6 = wb.create_sheet("Prédictions")
        ws6.append(["Année", "SNN", "CS", "PS", "GR", "NRV", "DWL", "EC"])
        _style_header(ws6, 1, 8)
        for p in predictions["projections"]:
            ws6.append([p["year"], p["snn"], p["cs"], p["ps"],
                         p["gr"], p["nrv"], p["dwl"], p["ec"]])
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws6.column_dimensions[col].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
